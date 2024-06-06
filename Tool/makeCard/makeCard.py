from openpyxl import load_workbook
from collections import deque

CARD_WIDTH = 3
CARD_HEIGHT = 6
MAX_CARD_PER_SHEET = 12
MAX_CARD_ROW_PER_SHEET = 4
MAX_CARD_COL_PER_SHEET = 3

INPUT_XL = load_workbook('cardUI.xlsx')

ORIGINAL_SHEET = INPUT_XL['Original']

STUDENT_INFO = INPUT_XL['Student']
CITIZEN_INFO = INPUT_XL['Citizen']
MAGIC_INFO = INPUT_XL['Magic']
TOKEN_INFO = INPUT_XL['Token']

DECK_INFO = INPUT_XL['Deck']


def main():
    global INPUT_XL, ORIGINAL_SHEET, STUDENT_INFO, CITIZEN_INFO, MAGIC_INFO, TOKEN_INFO, DECK_INFO

    page = 1
    deck_idx = 2
    current_card_amount = float("inf")
    current_sheet = None
    deck_cards_list = get_deck_info(DECK_INFO, 1)

    is_Finish = False

    while not is_Finish:
        current_sheet = copy_sheet(ORIGINAL_SHEET, "Instance"+str(page))

        for row in range(MAX_CARD_ROW_PER_SHEET):
            for col in range(MAX_CARD_COL_PER_SHEET):
                try:
                    card_info = deck_cards_list.pop()
                except:
                    is_Finish = True
                    break
                
                if card_info != None:
                    write_card(current_sheet, card_info, (1+6*row, 1+3*col))

                    if len(deck_cards_list) == 0:
                        is_Finish = True

                if is_Finish:
                    break

        if is_Finish:
            break

        page += 1
    INPUT_XL.save('output.xlsx')







def write_card(target_sheet, card_info, start_idx):
    offsets = {"Cost":(0,0), "Name":(0,1), "Effect":(1,0), "WP":(1,2), "Range":(2,2), "Club":(3,2), "Type":(4,2), "ATK":(5,0), "HP":(5,2)}

    #for key in card_info.keys():
    #    offset = offsets[key]
    #    
    #    target_sheet.cell(row=start_idx[0]+offset[0], column=start_idx[1]+offset[1]).value = card_info[key]

    for key in offsets.keys():
        try:
            target_sheet.cell(row=start_idx[0]+offsets[key][0], column=start_idx[1]+offsets[key][1]).value = card_info[key]
        except KeyError:
            target_sheet.cell(row=start_idx[0]+offsets[key][0], column=start_idx[1]+offsets[key][1]).value = ""
    return


def copy_sheet(target, name):
    global INPUT_XL

    copied_sheet = INPUT_XL.copy_worksheet(target)
    copied_sheet.title = name

    return copied_sheet

def find_card(cards_list, name):
    current_idx = 2
    while True:
        card_info = get_column_values_with_keys(cards_list, 1, current_idx)

        if len(card_info) == 0:
            return None

        if card_info["Name"] == name:
            return card_info
        
        current_idx += 1
        

def get_card_info(sheet, key_column, target_column):
    return get_column_values_with_keys(sheet, key_column, target_column)

def get_deck_info(sheet, key_column):
    global STUDENT_INFO, CITIZEN_INFO, MAGIC_INFO, TOKEN_INFO

    result = deque()
    current_idx = 2

    while True:
        deck_card = get_column_values_with_keys(sheet, key_column, current_idx)
        
        if len(deck_card) == 0:
            break
        if deck_card["Name"] != "-":
            if deck_card["Type"] == "학생":
                card_info = find_card(STUDENT_INFO, deck_card["Name"])
            elif deck_card["Type"] == "시민":
                card_info = find_card(CITIZEN_INFO, deck_card["Name"])
            elif deck_card["Type"] == "마법":
                card_info = find_card(MAGIC_INFO, deck_card["Name"])
            elif deck_card["Type"] == "토큰":
                card_info = find_card(TOKEN_INFO, deck_card["Name"])
            
            for _ in range(int(deck_card["Amount"])):
                result.append(card_info)

        current_idx += 1
    return result

def get_column_values_with_keys(sheet, key_column, target_column):
    data = {}
    current_col = 1
    
    for col in range(1, 10):
        if sheet.cell(row = target_column, column = col).value is not None:
            data[sheet.cell(row = key_column, column = col).value] = sheet.cell(row = target_column, column = col).value

    return data



main()
print("완료")
